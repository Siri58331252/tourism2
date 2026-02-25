from flask import Flask, render_template, redirect, request, send_file
from flask_login import (
    LoginManager, login_user, logout_user,
    login_required, current_user
)
from sqlalchemy import or_
from models import (
    db, User, Place, SystemLog,
    Category, Province, Recommend, Comment, PlaceImage
)
from werkzeug.utils import secure_filename
from flask import url_for
from datetime import datetime, timedelta

import csv
import io
import os
from datetime import datetime, timedelta
from flask import request, redirect, url_for, flash
from werkzeug.utils import secure_filename
import os
from sqlalchemy import or_
UPLOAD_FOLDER = 'static/uploads'
from flask_login import login_required, current_user

# ==================================================
# APP CONFIG
# ==================================================
app = Flask(__name__)
app.config['SECRET_KEY'] = 'secret'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///tourism.db'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
db.init_app(app)

# ==================================================
# CREATE DB + DEFAULT ADMIN
# ==================================================
with app.app_context():
    db.create_all()

    # ===== สร้าง sample users ทั้ง 3 role =====
    if not User.query.filter_by(username='admin').first():
        admin = User(
            username='admin',
            email='admin@tourism.com',
            password='admin123',
            role='admin',
            is_active=True
        )
        db.session.add(admin)
        print('✔ Admin user created')

    if not User.query.filter_by(username='moderator').first():
        moderator = User(
            username='moderator',
            email='moderator@tourism.com',
            password='moderator123',
            role='moderator',
            is_active=True
        )
        db.session.add(moderator)
        print('✔ Moderator user created')

    if not User.query.filter_by(username='member').first():
        member = User(
            username='member',
            email='member@tourism.com',
            password='member123',
            role='member',
            is_active=True
        )
        db.session.add(member)
        print('✔ Member user created')

    db.session.commit()



# ==================================================
# LOGIN MANAGER
# ==================================================
login_manager = LoginManager(app)
login_manager.login_view = 'login'


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# ==================================================
# PUBLIC PAGES
# ==================================================

# ---------- หน้าแรก (เฉพาะสถานที่ที่มีรีวิว) ----------
@app.route('/')

def index():
    keyword = request.args.get('q', '').strip()
    province = request.args.get('province')
    category = request.args.get('category')

    provinces = [p.name for p in Province.query.order_by(Province.name).all()]
    categories = [c.name for c in Category.query.order_by(Category.name).all()]

    query = (
        Place.query
        .filter(Place.approved == True)
        .join(Province)
        .join(Category)
    )

    if keyword:
        query = query.filter(
            or_(
                Place.name.contains(keyword),
                Province.name.contains(keyword),
                Category.name.contains(keyword)
            )
        )

    if province:
        query = query.filter(Province.name == province)

    if category:
        query = query.filter(Category.name == category)

    places = query.all()

    return render_template(
        'index.html',
        province=provinces,
        category=categories,
        places=places,
        keyword=keyword,
        selected_province=province,
        selected_category=category
    )

    

# ---------- รายละเอียดสถานที่ ----------
@app.route('/place/<int:id>')
def place_detail(id):
    place = Place.query.get_or_404(id)

    recommends = Recommend.query.filter_by(
        place_id=id,
        approved=True
    ).all()

    return render_template(
        'place_detail.html',
        place=place,
        recommends=recommends
    )



# ==================================================
# AUTH
# ==================================================

@app.route('/register', methods=['GET', 'POST'])
def register():
    error = None

    if request.method == 'POST':
        if User.query.filter_by(username=request.form['username']).first():
            error = 'ชื่อผู้ใช้นี้ถูกใช้แล้ว'
        else:
            user = User(
                username=request.form['username'],
                password=request.form['password'],
                email=request.form.get('email', '')
            )
            db.session.add(user)  
            db.session.commit()
            return redirect('/login')

    return render_template('register.html', error=error)


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None

    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        
        # ค้นหาผู้ใช้โดยใช้ email
        user = User.query.filter_by(email=email).first()

        if user and user.password == password:
            login_user(user)

            db.session.add(SystemLog(
                user_id=user.id,
                action='login',
                detail='User logged in',
                ip_address=request.remote_addr
            ))
            db.session.commit()

            return redirect('/')
        else:
            error = 'อีเมล หรือรหัสผ่านไม่ถูกต้อง'

    return render_template('login.html', error=error)

@app.route('/admin/recommend/<int:id>/approve')
@login_required
def approve_recommend(id):
    if current_user.role != 'admin':
        return redirect('/')

    rec = Recommend.query.get_or_404(id)
    rec.approved = True
    db.session.commit()

    return redirect(f'/place/{rec.place_id}')

@app.route('/users')
def admin_users():
    if current_user.role != 'admin':
        return redirect('/')
    
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    users = User.query.paginate(page=page, per_page=per_page)
    
    return render_template('users.html', users=users)


@app.route('/users/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_user(id):
    if current_user.role != 'admin':
        return redirect('/')
    
    user = User.query.get_or_404(id)

    if request.method == 'POST':
        # --- เพิ่มให้รองรับการแก้ไขทุกฟิลด์ ---
        user.username = request.form.get('username')
        user.email = request.form.get('email')
        user.role = request.form.get('role')
        user.is_active = request.form.get('is_active') == 'on'
        
        db.session.commit()
        
        # บันทึก Log การแก้ไขของ Admin
        db.session.add(SystemLog(
            user_id=current_user.id,
            action='edit_user',
            detail=f'Admin edited user: {user.username}',
            ip_address=request.remote_addr
        ))
        db.session.commit()
        
        return redirect('/users')

    # ตรวจสอบชื่อไฟล์ template ให้ตรงกับที่คุณมี (ในโค้ดคุณเขียน /user_edit.html)
    return render_template('user_edit.html', user=user)


@app.route('/logout')
@login_required
def logout():
    db.session.add(SystemLog(
        user_id=current_user.id,
        action='logout',
        detail='User logged out',
        ip_address=request.remote_addr
    ))
    db.session.commit()

    logout_user()
    return redirect('/login')

# ==================================================
# PLACE CRUD
# ==================================================


@app.route('/add_place', methods=['GET', 'POST'])
@login_required
def add_place():

    if request.method == 'POST':
        files = request.files.getlist('images')

        if not files or all(f.filename == '' for f in files):
            return 'กรุณาเลือกรูปภาพ'

        filenames = []
        for file in files[:10]:  # จำกัดเพียง 10 รูป
            if file and file.filename != '':
                filename = secure_filename(file.filename)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                filenames.append(filename)

        place = Place(
            name=request.form['name'],
            description=request.form['description'],
            image=filenames[0] if filenames else None,  # รูปแรกเป็นรูปหลัก
            category_id=request.form['category_id'],
            province_id=request.form['province_id'],
            user_id=current_user.id
        )

        db.session.add(place)
        db.session.flush()  # รับ place.id ก่อนสร้าง PlaceImage
        
        # เพิ่มรูปภาพลงใน PlaceImage
        for filename in filenames:
            place_image = PlaceImage(
                place_id=place.id,
                image_path=filename
            )
            db.session.add(place_image)
        
        db.session.commit()
        return redirect('/')

    return render_template(
        'add_place.html',
        categories=Category.query.all(),
        provinces=Province.query.all()
    )

@app.route('/place/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_place(id):
    place = Place.query.get_or_404(id)

    if current_user.role != 'admin' and place.user_id != current_user.id:
        return redirect('/')

    categories = Category.query.all()
    provinces = Province.query.all()

    if request.method == 'POST':
        place.name = request.form['name']
        place.description = request.form['description']
        place.category_id = request.form['category_id']
        place.province_id = request.form['province_id']
        place.approved = False

        # จัดการรูปภาพใหม่
        files = request.files.getlist('images')
        if files and any(f.filename != '' for f in files):
            # ลบรูปเก่า
            for img in place.images:
                db.session.delete(img)
            
            filenames = []
            for file in files[:10]:  # จำกัดเพียง 10 รูป
                if file and file.filename != '':
                    filename = secure_filename(file.filename)
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    filenames.append(filename)
            
            place.image = filenames[0] if filenames else None
            
            # เพิ่มรูปใหม่
            for filename in filenames:
                place_image = PlaceImage(
                    place_id=place.id,
                    image_path=filename
                )
                db.session.add(place_image)

        db.session.commit()

        db.session.add(SystemLog(
            user_id=current_user.id,
            action='edit_place',
            detail=f'Edited place: {place.name}',
            ip_address=request.remote_addr
        ))
        db.session.commit()

        return redirect(f'/place/{id}')

    return render_template(
        'edit_place.html',
        place=place,
        categories=categories,
        provinces=provinces
    )


@app.route('/place/<int:id>/delete', methods=['POST'])
@login_required
def delete_place(id):
    place = Place.query.get_or_404(id)

    if current_user.role != 'admin' and place.user_id != current_user.id:
        return redirect('/')

    db.session.add(SystemLog(
        user_id=current_user.id,
        action='delete_place',
        detail=f'Deleted place: {place.name}',
        ip_address=request.remote_addr
    ))

    db.session.delete(place)
    db.session.commit()

    return redirect('/admin/manage')

# ==================================================
# ADMIN
# ==================================================

@app.route('/admin/manage')
@login_required
def admin_manage():
    if current_user.role != 'admin':
        return redirect('/')

    return render_template(
        'manage_place.html',
        categories=Category.query.all(),
        provinces=Province.query.all(),
        places=Place.query.all()
    )


@app.route('/admin/category', methods=['POST'])
@login_required
def add_category():
    if current_user.role == 'admin':
        db.session.add(Category(name=request.form['name']))
        db.session.commit()
    return redirect('/admin/manage')


@app.route('/admin/category/<int:id>/edit', methods=['POST'])
@login_required
def edit_category(id):
    if current_user.role != 'admin':
        return redirect('/')
    
    category = Category.query.get_or_404(id)
    category.name = request.form['name']
    db.session.commit()
    return redirect('/admin/manage')


@app.route('/admin/category/<int:id>/delete', methods=['POST'])
@login_required
def delete_category(id):
    if current_user.role != 'admin':
        return redirect('/')
    
    category = Category.query.get_or_404(id)
    db.session.delete(category)
    db.session.commit()
    return redirect('/admin/manage')


@app.route('/admin/province', methods=['POST'])
@login_required
def add_province():
    if current_user.role == 'admin':
        db.session.add(Province(name=request.form['name']))
        db.session.commit()
    return redirect('/admin/manage')


@app.route('/admin/province/<int:id>/edit', methods=['POST'])
@login_required
def edit_province(id):
    if current_user.role != 'admin':
        return redirect('/')
    
    province = Province.query.get_or_404(id)
    province.name = request.form['name']
    db.session.commit()
    return redirect('/admin/manage')


@app.route('/admin/province/<int:id>/delete', methods=['POST'])
@login_required
def delete_province(id):
    if current_user.role != 'admin':
        return redirect('/')
    
    province = Province.query.get_or_404(id)
    db.session.delete(province)
    db.session.commit()
    return redirect('/admin/manage')


@app.route('/approve/<int:id>')
@login_required
def approve_place(id):
    if current_user.role != 'admin':
        return redirect('/')

    place = Place.query.get_or_404(id)
    place.approved = True
    db.session.commit()

    db.session.add(SystemLog(
        user_id=current_user.id,
        action='approve_place',
        detail=f'Approved place: {place.name}',
        ip_address=request.remote_addr
    ))
    db.session.commit()

    return redirect('/admin/manage')



@app.route("/place/<int:place_id>/comment", methods=["POST"])
@login_required
def add_comment(place_id):
    content = request.form['content']

    comment = Comment(
        user_id=current_user.id,
        place_id=place_id,
        content=content
    )

    db.session.add(comment)
    db.session.commit()

    return redirect(url_for('place_detail', id=place_id))



@app.route('/admin/place/<int:id>/recommend', methods=['POST'])
@login_required
def add_recommend(id):
    if current_user.role != 'admin':
        return redirect('/')

    db.session.add(Recommend(
        place_id=id,
        title=request.form['title'],
        content=request.form['content'],
        approved=False   # ✅ เพิ่มบรรทัดนี้
    ))
    db.session.commit()

    return redirect(f'/place/{id}')

# ==================================================
# REPORT
# ==================================================

@app.route('/report')
@login_required
def report():
    if current_user.role != 'admin':
        return redirect('/')

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Name', 'Province', 'Category', 'Approved'])

    for p in Place.query.all():
        writer.writerow([
            p.name,
            p.province.name,
            p.category.name,
            p.approved
        ])

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode()),
        mimetype='text/csv',
        as_attachment=True,
        download_name='tourism_report.csv'
    )




@app.route('/my-places')
@login_required
def my_places():
    places = Place.query.filter_by(user_id=current_user.id) \
                         .order_by(Place.id.desc()) \
                         .all()

    return render_template('my_places.html', places=places)


@app.route('/my-recommends')
@login_required
def my_recommends():
    recommends = Recommend.query.filter(
        Recommend.place.has(Place.user_id == current_user.id)
    ).order_by(Recommend.created_at.desc()).all()
    return render_template('my_recommends.html', recommends=recommends)


@app.route('/admin')
@login_required
def admin():
    if current_user.role != 'admin':
        return redirect('/')
    
    total_places = Place.query.count()
    approved_places = Place.query.filter_by(approved=True).count()
    pending_places = Place.query.filter_by(approved=False).count()
    total_users = User.query.count()
    
    return render_template(
        'admin.html',
        total_places=total_places,
        approved_places=approved_places,
        pending_places=pending_places,
        total_users=total_users
    )


# ==================================================
# MODERATOR (ผู้ดูแลระบบ)
# ==================================================

@app.route('/moderator')
@login_required
def moderator():
    if current_user.role != 'moderator':
        return redirect('/')
    
    total_users = User.query.count()
    active_users = User.query.filter_by(is_active=True).count()
    inactive_users = User.query.filter_by(is_active=False).count()
    total_logs = SystemLog.query.count()
    
    return render_template(
        'moderator.html',
        total_users=total_users,
        active_users=active_users,
        inactive_users=inactive_users,
        total_logs=total_logs
    )


@app.route('/moderator/members')
@login_required
def moderator_members():
    if current_user.role != 'moderator':
        return redirect('/')
    
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    users = User.query.paginate(page=page, per_page=per_page)
    
    return render_template('moderator_members.html', users=users)


@app.route('/moderator/member/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def moderator_edit_member(id):
    if current_user.role != 'moderator':
        return redirect('/')
    
    user = User.query.get_or_404(id)
    
    if request.method == 'POST':
        # --- เพิ่มส่วนที่ดึงข้อมูลจากฟอร์มให้ครบทุกช่อง ---
        user.username = request.form.get('username')
        user.email = request.form.get('email')
        user.role = request.form.get('role')
        
        # สำหรับ checkbox ถ้าไม่ได้ติ๊กมันจะส่ง None มา เราจึงเช็คว่าเป็น 'on' หรือไม่
        user.is_active = request.form.get('is_active') == 'on'
        
        # บันทึกลงฐานข้อมูล
        db.session.commit()
        
        # บันทึก Log
        db.session.add(SystemLog(
            user_id=current_user.id,
            action='edit_member',
            detail=f'Moderator edited member: {user.username} (Role: {user.role})',
            ip_address=request.remote_addr
        ))
        db.session.commit()
        
        return redirect('/moderator/members')
    
    return render_template('moderator_edit_member.html', user=user)


@app.route('/moderator/logs')
@login_required
def moderator_logs():
    if current_user.role != 'moderator':
        return redirect('/')
    
    from datetime import datetime, timedelta
    
    # รับช่วงเวลาจาก query parameters
    days = request.args.get('days', 7, type=int)
    action = request.args.get('action', '', type=str)
    
    # คำนวณวันที่เริ่มต้น
    start_date = datetime.utcnow() - timedelta(days=days)
    
    # Query logs ตามช่วงเวลา
    query = SystemLog.query.filter(SystemLog.created_at >= start_date)
    
    if action:
        query = query.filter_by(action=action)
    
    logs = query.order_by(SystemLog.created_at.desc()).all()
    
    # สถิติตามประเภท
    actions = db.session.query(SystemLog.action, db.func.count(SystemLog.id)) \
        .filter(SystemLog.created_at >= start_date) \
        .group_by(SystemLog.action).all()
    
    return render_template(
        'moderator_logs.html',
        logs=logs,
        days=days,
        action=action,
        actions=actions
    )


@app.route('/moderator/logs/export-csv')
@login_required
def export_logs_csv():
    if current_user.role != 'moderator':
        return redirect('/')
    
    days = request.args.get('days', 7, type=int)
    action = request.args.get('action', '', type=str)
    
    # คำนวณวันที่เริ่มต้น
    start_date = datetime.utcnow() - timedelta(days=days)
    
    # Query logs ตามช่วงเวลา
    query = SystemLog.query.filter(SystemLog.created_at >= start_date)
    
    if action:
        query = query.filter_by(action=action)
    
    logs = query.order_by(SystemLog.created_at.desc()).all()
    
    # สร้าง CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['วันเวลา', 'ผู้ใช้', 'การทำงาน', 'รายละเอียด', 'IP Address'])
    
    for log in logs:
        writer.writerow([
            log.created_at.strftime('%d/%m/%Y %H:%M:%S') if log.created_at else '',
            log.user.username if log.user else 'ลบแล้ว',
            log.action,
            log.detail if log.detail else '',
            log.ip_address if log.ip_address else ''
        ])
    
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv; charset=utf-8',
        as_attachment=True,
        download_name=f'system_logs_{datetime.now().strftime("%d%m%Y_%H%M%S")}.csv'
    )


@app.route('/moderator/logs/export-excel')
@login_required
def export_logs_excel():
    if current_user.role != 'moderator':
        return redirect('/')
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return 'กรุณาติดตั้ง openpyxl: pip install openpyxl', 400
    
    days = request.args.get('days', 7, type=int)
    action = request.args.get('action', '', type=str)
    
    # คำนวณวันที่เริ่มต้น
    start_date = datetime.utcnow() - timedelta(days=days)
    
    # Query logs ตามช่วงเวลา
    query = SystemLog.query.filter(SystemLog.created_at >= start_date)
    
    if action:
        query = query.filter_by(action=action)
    
    logs = query.order_by(SystemLog.created_at.desc()).all()
    
    # สร้าง Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "System Logs"
    
    # ตั้งค่าส่วนหัว
    headers = ['วันเวลา', 'ผู้ใช้', 'การทำงาน', 'รายละเอียด', 'IP Address']
    ws.append(headers)
    
    # สไตล์ส่วนหัว
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # เพิ่มข้อมูล
    for log in logs:
        ws.append([
            log.created_at.strftime('%d/%m/%Y %H:%M:%S') if log.created_at else '',
            log.user.username if log.user else 'ลบแล้ว',
            log.action,
            log.detail if log.detail else '',
            log.ip_address if log.ip_address else ''
        ])
    
    # ตั้งความกว้างของคอลัมน์
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    
    # เพิ่มเส้นขอบให้ข้อมูล
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # บันทึกเป็น bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'system_logs_{datetime.now().strftime("%d%m%Y_%H%M%S")}.xlsx'
    )


@app.route('/moderator/logs/export-pdf')
@login_required
def export_logs_pdf():
    if current_user.role != 'moderator':
        return redirect('/')
    
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return 'กรุณาติดตั้ง reportlab: pip install reportlab', 400
    
    days = request.args.get('days', 7, type=int)
    action = request.args.get('action', '', type=str)
    
    # คำนวณวันที่เริ่มต้น
    start_date = datetime.utcnow() - timedelta(days=days)
    
    # Query logs ตามช่วงเวลา
    query = SystemLog.query.filter(SystemLog.created_at >= start_date)
    
    if action:
        query = query.filter_by(action=action)
    
    logs = query.order_by(SystemLog.created_at.desc()).all()
    
    # สร้าง PDF
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # เนื้อหา
    elements = []
    
    # ชื่อเอกสาร
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=12,
        alignment=1  # center
    )
    title = Paragraph(f'ประวัติการใช้งานระบบ ({days} วัน)', title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # ตาราง
    data = [['วันเวลา', 'ผู้ใช้', 'การทำงาน', 'รายละเอียด', 'IP Address']]
    
    for log in logs[:100]:  # จำกัดไม่เกิน 100 แถว เพื่อไม่ให้ PDF ใหญ่เกินไป
        data.append([
            log.created_at.strftime('%d/%m/%Y %H:%M') if log.created_at else '',
            log.user.username if log.user else 'ลบแล้ว',
            log.action,
            (log.detail if log.detail else '')[:30],  # จำกัดความยาว
            log.ip_address if log.ip_address else ''
        ])
    
    table = Table(data, colWidths=[1.3*inch, 1*inch, 1*inch, 2*inch, 1.2*inch])
    
    # สไตล์ตาราง
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
    ]))
    
    elements.append(table)
    
    # สร้าง PDF
    doc.build(elements)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'system_logs_{datetime.now().strftime("%d%m%Y_%H%M%S")}.pdf'
    )


# ==================================================
# RUN
# ==================================================
if __name__ == '__main__':
    app.run(debug=True)
