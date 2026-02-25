import os
import io
import csv
from datetime import datetime, timedelta

from flask import Flask, render_template, redirect, request, send_file, url_for, flash
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from sqlalchemy import or_
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash # ✅ เพิ่มระบบเข้ารหัสรหัสผ่าน
from flask_migrate import Migrate
from models import (
    db, User, Place, SystemLog,
    Category, Province, Recommend, Comment, PlaceImage
)

# ==================================================
# APP CONFIG
# ==================================================
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'} # ✅ จำกัดนามสกุลไฟล์รูปภาพ

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your_strong_secret_key_here' # ⚠️ ควรเปลี่ยนเป็นคีย์ที่เดายากใน Production
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///tourism.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024 # ✅ จำกัดขนาดไฟล์อัปโหลดสูงสุดที่ 16MB

db.init_app(app)
migrate = Migrate(app, db)

# ตรวจสอบนามสกุลไฟล์
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# สร้างโฟลเดอร์สำหรับอัปโหลดถ้ายังไม่มี
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ==================================================
# CREATE DB + DEFAULT ADMIN
# ==================================================
with app.app_context():
    db.create_all()

    if Province.query.count() == 0:
        thai_provinces = [
            "กรุงเทพมหานคร","กระบี่","กาญจนบุรี","กาฬสินธุ์","กำแพงเพชร",
            "ขอนแก่น","จันทบุรี","ฉะเชิงเทรา","ชลบุรี","ชัยนาท",
            "ชัยภูมิ","ชุมพร","เชียงราย","เชียงใหม่","ตรัง",
            "ตราด","ตาก","นครนายก","นครปฐม","นครพนม",
            "นครราชสีมา","นครศรีธรรมราช","นครสวรรค์","นนทบุรี","นราธิวาส",
            "น่าน","บึงกาฬ","บุรีรัมย์","ปทุมธานี","ประจวบคีรีขันธ์",
            "ปราจีนบุรี","ปัตตานี","พระนครศรีอยุธยา","พังงา","พัทลุง",
            "พิจิตร","พิษณุโลก","เพชรบุรี","เพชรบูรณ์","แพร่",
            "พะเยา","ภูเก็ต","มหาสารคาม","มุกดาหาร","แม่ฮ่องสอน",
            "ยะลา","ยโสธร","ร้อยเอ็ด","ระนอง","ระยอง",
            "ราชบุรี","ลพบุรี","ลำปาง","ลำพูน","เลย",
            "ศรีสะเกษ","สกลนคร","สงขลา","สตูล","สมุทรปราการ",
            "สมุทรสงคราม","สมุทรสาคร","สระแก้ว","สระบุรี","สิงห์บุรี",
            "สุโขทัย","สุพรรณบุรี","สุราษฎร์ธานี","สุรินทร์","หนองคาย",
            "หนองบัวลำภู","อ่างทอง","อำนาจเจริญ","อุดรธานี","อุตรดิตถ์",
            "อุทัยธานี","อุบลราชธานี"
        ]

        for province_name in thai_provinces:
            db.session.add(Province(name=province_name))
        db.session.commit()

    # ✅ เข้ารหัสรหัสผ่านก่อนบันทึกลง Database
    if not User.query.filter_by(username='admin').first():
        admin = User(
            username='admin',
            email='admin@gmail.com',
            password=generate_password_hash('admin'),
            role='admin',
            is_active=True
        )
        db.session.add(admin)
        print('✔ Admin user created')

    if not User.query.filter_by(username='moderator').first():
        moderator = User(
            username='moderator',
            email='moderator@gmail.com',
            password=generate_password_hash('moderator123'),
            role='moderator',
            is_active=True
        )
        db.session.add(moderator)
        print('✔ Moderator user created')

    if not User.query.filter_by(username='member').first():
        member = User(
            username='member',
            email='member@gmail.com',
            password=generate_password_hash('member123'),
            role='member',
            is_active=True
        )
        db.session.add(member)
        print('✔ Member user created')

    db.session.commit()

# ==================================================
# LOGIN MANAGER
# ==================================================
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = "กรุณาเข้าสู่ระบบก่อนเข้าใช้งาน"
login_manager.login_message_category = "warning"

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# ==================================================
# PUBLIC PAGES
# ==================================================

@app.route('/')
def index():
    keyword = request.args.get('q', '').strip()
    province = request.args.get('province')
    category = request.args.get('category')

    provinces = [p.name for p in Province.query.order_by(Province.name).all()]
    categories = [c.name for c in Category.query.order_by(Category.name).all()]

    query = Place.query.filter(Place.approved == True).join(Province).join(Category)

    if keyword:
        query = query.filter(
            or_(
                Place.name.ilike(f'%{keyword}%'),
                Province.name.ilike(f'%{keyword}%'),
                Category.name.ilike(f'%{keyword}%')
            )
        )

    if province:
        query = query.filter(Province.name == province)

    if category:
        query = query.filter(Category.name == category)

    places = query.all()

    return render_template(
        'index.html',
        province=provinces,
        category=categories,
        places=places,
        keyword=keyword,
        selected_province=province,
        selected_category=category
    )

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        new_username = request.form.get('username')
        new_surname = request.form.get('surname')  # ✅ รับค่านามสกุลจากฟอร์ม
        email = request.form.get('email')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        # 1. ตรวจสอบชื่อผู้ใช้
        if new_username != current_user.username:
            user_exists = User.query.filter_by(username=new_username).first()
            if user_exists:
                flash("ชื่อผู้ใช้งานนี้ถูกใช้ไปแล้ว", "danger")
                return redirect(url_for('profile'))
            current_user.username = new_username

        # 2. อัปเดตนามสกุลและอีเมล
        current_user.surname = new_surname  # ✅ บันทึกนามสกุลลง Database
        current_user.email = email

        # 3. ตรวจสอบการเปลี่ยนรหัสผ่าน (เหมือนเดิม)
        if new_password:
            if new_password == confirm_password:
                current_user.password = generate_password_hash(new_password)
            else:
                flash("รหัสผ่านไม่ตรงกัน", "danger")
                return redirect(url_for('profile'))

        try:
            db.session.commit()
            flash("อัปเดตข้อมูล ชื่อ-นามสกุล เรียบร้อยแล้ว!", "success")
        except Exception as e:
            db.session.rollback()
            flash("เกิดข้อผิดพลาดในการบันทึกข้อมูล", "danger")
            
        return redirect(url_for('profile'))

    return render_template('profile.html')

@app.route('/place/<int:id>')
def place_detail(id):
    place = Place.query.get_or_404(id)
    recommends = Recommend.query.filter_by(place_id=id, approved=True).all()
    return render_template('place_detail.html', place=place, recommends=recommends)

# ==================================================
# AUTH & USER MANAGEMENT
# ==================================================

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        surname = request.form.get('surname') # รับค่านามสกุล
        email = request.form.get('email')
        password = request.form.get('password')

        # 1. ตรวจสอบว่ามีชื่อผู้ใช้นี้ในระบบหรือยัง
        user_exists = User.query.filter_by(username=username).first()
        if user_exists:
            return render_template('register.html', error="ชื่อผู้ใช้งานนี้มีผู้ใช้แล้ว")

        # 2. แฮชรหัสผ่าน (Password Hashing)
        hashed_password = generate_password_hash(password)

        # 3. สร้าง User ใหม่พร้อมเก็บนามสกุล
        new_user = User(
            username=username,
            surname=surname,  # เก็บลงคอลัมน์ surname ที่เราสร้างไว้
            email=email,
            password=hashed_password,
            role='member'     # กำหนดค่าเริ่มต้นเป็น member
        )

        try:
            db.session.add(new_user)
            db.session.commit()
            # สมัครเสร็จแล้วส่งไปหน้า Login
            flash("สมัครสมาชิกสำเร็จ! กรุณาเข้าสู่ระบบ", "success")
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            return render_template('register.html', error="เกิดข้อผิดพลาดในการบันทึกข้อมูล")

    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        
        user = User.query.filter_by(email=email).first()

        # ✅ ตรวจสอบรหัสผ่านที่เข้ารหัสไว้
        if user and check_password_hash(user.password, password):
            if not user.is_active:
                error = 'บัญชีของคุณถูกระงับการใช้งาน'
                return render_template('login.html', error=error)

            login_user(user)
            db.session.add(SystemLog(
                user_id=user.id,
                action='login',
                detail='User logged in',
                ip_address=request.remote_addr
            ))
            db.session.commit()
            return redirect(url_for('index'))
        else:
            error = 'อีเมล หรือรหัสผ่านไม่ถูกต้อง'

    return render_template('login.html', error=error)

@app.route('/logout')
@login_required
def logout():
    db.session.add(SystemLog(
        user_id=current_user.id,
        action='logout',
        detail='User logged out',
        ip_address=request.remote_addr
    ))
    db.session.commit()
    logout_user()
    return redirect(url_for('login'))

@app.route('/users')
@login_required
def admin_users():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    
    page = request.args.get('page', 1, type=int)
    users = User.query.paginate(page=page, per_page=20)
    return render_template('users.html', users=users)

@app.route('/users/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_user(id):
    # 1. ตรวจสอบสิทธิ์ Admin
    if current_user.role != 'admin':
        flash('คุณไม่มีสิทธิ์เข้าถึงส่วนนี้', 'error')
        return redirect(url_for('index'))
    
    # 2. ดึงข้อมูล User (ถ้าไม่เจอจะขึ้น 404)
    user = User.query.get_or_404(id)

    if request.method == 'POST':
        # ใช้ .get(..., '') เพื่อป้องกันกรณีฟิลด์หาย จะได้ไม่ Error
        new_username = request.form.get('username', '').strip()
        new_email = request.form.get('email', '').strip()
        new_surname = request.form.get('surname', '').strip()

        # 3. เช็คความซ้ำซ้อน
        existing_user = User.query.filter(
            User.id != id, 
            or_(User.username == new_username, User.email == new_email)
        ).first()

        if existing_user:
            if existing_user.username == new_username:
                flash('ชื่อผู้ใช้นี้ถูกใช้โดยบัญชีอื่นแล้ว', 'error')
            else:
                flash('อีเมลนี้ถูกใช้โดยบัญชีอื่นแล้ว', 'error')
            # ส่งกลับไปหน้าเดิมพร้อมส่ง user ตัวเดิมกลับไปด้วย (ถ้าจำเป็น)
            return render_template('user_edit.html', user=user)

        # 4. อัปเดตข้อมูล
        user.username = new_username
        user.surname = new_surname
        user.email = new_email
        user.role = request.form.get('role')
        user.is_active = request.form.get('is_active') == 'on'
        
        # จัดการรหัสผ่าน
        new_password = request.form.get('password')
        if new_password and len(new_password.strip()) > 0:
            user.password = generate_password_hash(new_password)
        
        try:
            db.session.add(SystemLog(
                user_id=current_user.id,
                action='edit_user',
                detail=f'Admin edited user: {user.username}',
                ip_address=request.remote_addr
            ))
            db.session.commit()
            flash(f'บันทึกข้อมูลของ {user.username} เรียบร้อยแล้ว', 'success')
            # ตรวจสอบชื่อฟังก์ชันใน url_for ให้ตรงกับหน้าตารางผู้ใช้ของคุณ
            return redirect(url_for('admin_users')) 
        except Exception as e:
            db.session.rollback()
            flash('เกิดข้อผิดพลาดในการบันทึกข้อมูล', 'error')
            print(f"Debug Error: {e}")

    # 5. ส่งตัวแปร user ไปที่หน้า Template (สำคัญมาก!)
    return render_template('user_edit.html', user=user)

# ==================================================
# PLACE CRUD
# ==================================================

@app.route('/add_place', methods=['GET', 'POST'])
@login_required
def add_place():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        
        if Place.query.filter_by(name=name).first():
            flash(f'สถานที่ชื่อ "{name}" มีอยู่ในระบบแล้ว!', 'error')
            return redirect(request.url)

        lat_str = request.form.get('latitude', '').strip()
        lng_str = request.form.get('longitude', '').strip()
        # ✅ กัน Error แปลงค่าว่างเป็น float
        lat = float(lat_str) if lat_str else None
        lng = float(lng_str) if lng_str else None

        files = request.files.getlist('images')
        filenames = []
        
        for file in files[:10]:  
            if file and file.filename != '' and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                # เพิ่ม timestamp ป้องกันชื่อรูปซ้ำ
                unique_filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}"
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], unique_filename))
                filenames.append(unique_filename)

        if not filenames:
            flash('กรุณาเลือกรูปภาพอย่างน้อย 1 รูป (เฉพาะ png, jpg, jpeg, gif, webp)', 'error')
            return redirect(request.url)

        place = Place(
            name=name,
            description=request.form['description'],
            image=filenames[0], 
            category_id=request.form['category_id'],
            province_id=request.form['province_id'],
            user_id=current_user.id,
            latitude=lat,
            longitude=lng
        )

        db.session.add(place)
        db.session.flush() 
        
        for fname in filenames:
            db.session.add(PlaceImage(place_id=place.id, image_path=fname))
        
        db.session.commit()
        flash('เพิ่มสถานที่เรียบร้อย รอการอนุมัติ', 'success')
        return redirect(url_for('index'))

    return render_template(
        'add_place.html',
        categories=Category.query.all(),
        provinces=Province.query.all()
    )

@app.route('/place/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_place(id):
    place = Place.query.get_or_404(id)

    if current_user.role != 'admin' and place.user_id != current_user.id:
        return redirect(url_for('index'))

    if request.method == 'POST':
        new_name = request.form.get('name', '').strip()

        if Place.query.filter(Place.name == new_name, Place.id != id).first():
            flash(f'สถานที่ชื่อ "{new_name}" มีอยู่ในระบบแล้ว!', 'error')
            return redirect(request.url)

        lat_str = request.form.get('latitude', '').strip()
        lng_str = request.form.get('longitude', '').strip()

        place.name = new_name
        place.description = request.form['description']
        place.category_id = request.form['category_id']
        place.province_id = request.form['province_id']
        place.approved = False
        
        place.latitude = float(lat_str) if lat_str else place.latitude
        place.longitude = float(lng_str) if lng_str else place.longitude

        files = request.files.getlist('images')
        # เช็คว่ามีการอัปโหลดไฟล์ใหม่ที่ใช้ได้หรือไม่
        has_valid_files = any(f and f.filename != '' and allowed_file(f.filename) for f in files)

        if has_valid_files:
            # ลบรูปเก่าออกจากฐานข้อมูล (ถ้าต้องการลบไฟล์จริงออกจากโฟลเดอร์ด้วย ต้องใช้ os.remove)
            for img in place.images:
                db.session.delete(img)
            
            filenames = []
            for file in files[:10]:  
                if file and file.filename != '' and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    unique_filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}"
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], unique_filename))
                    filenames.append(unique_filename)
            
            if filenames:
                place.image = filenames[0]
                for fname in filenames:
                    db.session.add(PlaceImage(place_id=place.id, image_path=fname))

        db.session.add(SystemLog(
            user_id=current_user.id,
            action='edit_place',
            detail=f'Edited place: {place.name}',
            ip_address=request.remote_addr
        ))
        db.session.commit()

        flash('อัปเดตข้อมูลสถานที่เรียบร้อยแล้ว', 'success')
        return redirect(url_for('place_detail', id=id))

    return render_template(
        'edit_place.html', 
        place=place, 
        categories=Category.query.all(), 
        provinces=Province.query.all()
    )

@app.route('/place/<int:id>/delete', methods=['POST'])
@login_required
def delete_place(id):
    place = Place.query.get_or_404(id)

    if current_user.role != 'admin' and place.user_id != current_user.id:
        return redirect(url_for('index'))

    db.session.add(SystemLog(
        user_id=current_user.id,
        action='delete_place',
        detail=f'Deleted place: {place.name}',
        ip_address=request.remote_addr
    ))
    db.session.delete(place)
    db.session.commit()
    flash('ลบสถานที่เรียบร้อยแล้ว', 'success')
    return redirect(url_for('admin_manage'))

# ==================================================
# ADMIN MANAGE (Category, Province, Place Approval)
# ==================================================

@app.route('/admin/manage')
@login_required
def admin_manage():
    if current_user.role != 'admin':
        return redirect(url_for('index'))

    return render_template(
        'manage_place.html',
        categories=Category.query.all(),
        provinces=Province.query.all(),
        places=Place.query.all()
    )

@app.route('/admin/category', methods=['POST'])
@login_required
def add_category():
    if current_user.role != 'admin': return redirect(url_for('index'))
    name = request.form.get('name', '').strip()
    
    if Category.query.filter_by(name=name).first():
        flash(f'หมวดหมู่ "{name}" มีอยู่ในระบบแล้ว!', 'error')
    else:
        db.session.add(Category(name=name))
        db.session.commit()
        flash('เพิ่มหมวดหมู่สำเร็จ', 'success')
        
    return redirect(url_for('admin_manage'))

@app.route('/admin/category/<int:id>/edit', methods=['POST'])
@login_required
def edit_category(id):
    if current_user.role != 'admin': return redirect(url_for('index'))
    category = Category.query.get_or_404(id)
    new_name = request.form.get('name', '').strip()

    if Category.query.filter(Category.name == new_name, Category.id != id).first():
        flash(f'หมวดหมู่ "{new_name}" มีอยู่ในระบบแล้ว!', 'error')
    else:
        category.name = new_name
        db.session.commit()
        flash('แก้ไขหมวดหมู่สำเร็จ', 'success')

    return redirect(url_for('admin_manage'))

@app.route('/admin/category/<int:id>/delete', methods=['POST'])
@login_required
def delete_category(id):
    if current_user.role != 'admin': return redirect(url_for('index'))
    category = Category.query.get_or_404(id)
    db.session.delete(category)
    db.session.commit()
    return redirect(url_for('admin_manage'))

@app.route('/admin/province', methods=['POST'])
@login_required
def add_province():
    if current_user.role != 'admin': return redirect(url_for('index'))
    name = request.form.get('name', '').strip()
    
    if not name:
        flash('กรุณากรอกชื่อจังหวัด', 'warning')
        return redirect(url_for('admin_manage'))

    if Province.query.filter_by(name=name).first():
        flash(f'จังหวัด "{name}" มีอยู่ในระบบแล้ว!', 'error')
    else:
        try:
            db.session.add(Province(name=name))
            db.session.commit()
            flash(f'เพิ่มจังหวัด "{name}" เรียบร้อยแล้ว', 'success')
        except Exception:
            db.session.rollback()
            flash('เกิดข้อผิดพลาดในการบันทึกข้อมูล', 'error')
                
    return redirect(url_for('admin_manage'))

@app.route('/admin/province/<int:id>/edit', methods=['POST'])
@login_required
def edit_province(id):
    if current_user.role != 'admin': return redirect(url_for('index'))
    province = Province.query.get_or_404(id)
    new_name = request.form.get('name', '').strip()

    if Province.query.filter(Province.name == new_name, Province.id != id).first():
        flash(f'จังหวัด "{new_name}" มีอยู่ในระบบแล้ว!', 'error')
    else:
        province.name = new_name
        db.session.commit()
        flash('แก้ไขจังหวัดสำเร็จ', 'success')

    return redirect(url_for('admin_manage'))

@app.route('/admin/province/<int:id>/delete', methods=['POST'])
@login_required
def delete_province(id):
    if current_user.role != 'admin': return redirect(url_for('index'))
    province = Province.query.get_or_404(id)
    db.session.delete(province)
    db.session.commit()
    return redirect(url_for('admin_manage'))

@app.route('/approve/<int:id>')
@login_required
def approve_place(id):
    if current_user.role != 'admin': return redirect(url_for('index'))
    place = Place.query.get_or_404(id)
    place.approved = True
    
    db.session.add(SystemLog(
        user_id=current_user.id,
        action='approve_place',
        detail=f'Approved place: {place.name}',
        ip_address=request.remote_addr
    ))
    db.session.commit()
    flash('อนุมัติสถานที่สำเร็จ', 'success')
    return redirect(url_for('admin_manage'))

# ==================================================
# COMMENTS & RECOMMENDS
# ==================================================

@app.route("/place/<int:place_id>/comment", methods=["POST"])
@login_required
def add_comment(place_id):
    content = request.form.get('content', '').strip()
    if content:
        comment = Comment(
            user_id=current_user.id,
            place_id=place_id,
            content=content
        )
        db.session.add(comment)
        db.session.commit()
    return redirect(url_for('place_detail', id=place_id))

@app.route('/admin/place/<int:id>/recommend', methods=['POST'])
@login_required
def add_recommend(id):
    if current_user.role != 'admin': return redirect(url_for('index'))

    title = request.form.get('title', '').strip()
    
    if Recommend.query.filter_by(place_id=id, title=title).first():
        flash('คุณเคยแนะนำหัวข้อนี้สำหรับสถานที่นี้ไปแล้ว', 'error')
        return redirect(url_for('place_detail', id=id))

    db.session.add(Recommend(
        place_id=id,
        title=title,
        content=request.form.get('content', ''),
        approved=False
    ))
    db.session.commit()
    flash('เพิ่มคำแนะนำสำเร็จ รอการอนุมัติ', 'success')

    return redirect(url_for('place_detail', id=id))

@app.route('/admin/recommend/<int:id>/approve')
@login_required
def approve_recommend(id):
    if current_user.role != 'admin': return redirect(url_for('index'))
    rec = Recommend.query.get_or_404(id)
    rec.approved = True
    db.session.commit()
    return redirect(url_for('place_detail', id=rec.place_id))

# ==================================================
# USER DASHBOARD (My Places & Recommends)
# ==================================================

@app.route('/my-places')
@login_required
def my_places():
    places = Place.query.filter_by(user_id=current_user.id).order_by(Place.id.desc()).all()
    return render_template('my_places.html', places=places)

@app.route('/my-recommends')
@login_required
def my_recommends():
    recommends = Recommend.query.filter(
        Recommend.place.has(Place.user_id == current_user.id)
    ).order_by(Recommend.created_at.desc()).all()
    return render_template('my_recommends.html', recommends=recommends)

# ==================================================
# ADMIN DASHBOARD & REPORT
# ==================================================

@app.route('/admin')
@login_required
def admin():
    if current_user.role != 'admin': return redirect(url_for('index'))
    
    total_places = Place.query.count()
    approved_places = Place.query.filter_by(approved=True).count()
    pending_places = Place.query.filter_by(approved=False).count()
    total_users = User.query.count()
    
    return render_template(
        'admin.html',
        total_places=total_places,
        approved_places=approved_places,
        pending_places=pending_places,
        total_users=total_users
    )

@app.route('/report')
@login_required
def report():
    if current_user.role != 'admin': return redirect(url_for('index'))

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Name', 'Province', 'Category', 'Approved'])

    for p in Place.query.all():
        writer.writerow([
            p.name,
            p.province.name,
            p.category.name,
            p.approved
        ])

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name='tourism_report.csv'
    )

# ==================================================
# MODERATOR (ผู้ดูแลระบบ)
# ==================================================

@app.route('/moderator')
@login_required
def moderator():
    if current_user.role != 'moderator': return redirect(url_for('index'))
    
    total_users = User.query.count()
    active_users = User.query.filter_by(is_active=True).count()
    inactive_users = User.query.filter_by(is_active=False).count()
    total_logs = SystemLog.query.count()
    
    return render_template(
        'moderator.html',
        total_users=total_users,
        active_users=active_users,
        inactive_users=inactive_users,
        total_logs=total_logs
    )

@app.route('/moderator/members')
@login_required
def moderator_members():
    if current_user.role != 'moderator': return redirect(url_for('index'))
    page = request.args.get('page', 1, type=int)
    users = User.query.paginate(page=page, per_page=20)
    return render_template('moderator_members.html', users=users)

@app.route('/moderator/member/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def moderator_edit_member(id):
    if current_user.role != 'moderator': return redirect(url_for('index'))
    
    user = User.query.get_or_404(id)
    
    if request.method == 'POST':
        new_username = request.form.get('username', '').strip()
        new_email = request.form.get('email', '').strip()

        existing_user = User.query.filter(User.id != id, or_(User.username == new_username, User.email == new_email)).first()
        if existing_user:
            if existing_user.username == new_username:
                flash('ชื่อผู้ใช้นี้ถูกใช้โดยบัญชีอื่นแล้ว', 'error')
            else:
                flash('อีเมลนี้ถูกใช้โดยบัญชีอื่นแล้ว', 'error')
            return redirect(url_for('moderator_edit_member', id=id))

        user.username = new_username
        user.email = new_email
        user.role = request.form.get('role')
        user.is_active = request.form.get('is_active') == 'on'
        
        db.session.add(SystemLog(
            user_id=current_user.id,
            action='edit_member',
            detail=f'Moderator edited member: {user.username} (Role: {user.role})',
            ip_address=request.remote_addr
        ))
        db.session.commit()
        
        flash('อัปเดตข้อมูลผู้ใช้สำเร็จ', 'success')
        return redirect(url_for('moderator_members'))
    
    return render_template('moderator_edit_member.html', user=user)

@app.route('/moderator/logs')
@login_required
def moderator_logs():
    if current_user.role != 'moderator': return redirect(url_for('index'))
    
    days = request.args.get('days', 7, type=int)
    action = request.args.get('action', '', type=str)
    
    start_date = datetime.now() - timedelta(days=days)
    query = SystemLog.query.filter(SystemLog.created_at >= start_date)
    
    if action: query = query.filter_by(action=action)
    logs = query.order_by(SystemLog.created_at.desc()).all()
    
    actions = db.session.query(SystemLog.action, db.func.count(SystemLog.id)) \
        .filter(SystemLog.created_at >= start_date) \
        .group_by(SystemLog.action).all()
    
    return render_template('moderator_logs.html', logs=logs, days=days, action=action, actions=actions)

@app.route('/moderator/logs/export-csv')
@login_required
def export_logs_csv():
    if current_user.role != 'moderator': return redirect(url_for('index'))
    
    days = request.args.get('days', 7, type=int)
    action = request.args.get('action', '', type=str)
    
    start_date = datetime.now() - timedelta(days=days)
    query = SystemLog.query.filter(SystemLog.created_at >= start_date)
    if action: query = query.filter_by(action=action)
    logs = query.order_by(SystemLog.created_at.desc()).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['วันเวลา', 'ผู้ใช้', 'การทำงาน', 'รายละเอียด', 'IP Address'])
    
    for log in logs:
        writer.writerow([
            log.created_at.strftime('%d/%m/%Y %H:%M:%S') if log.created_at else '',
            log.user.username if log.user else 'ลบแล้ว',
            log.action,
            log.detail if log.detail else '',
            log.ip_address if log.ip_address else ''
        ])
    
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'system_logs_{datetime.now().strftime("%d%m%Y_%H%M%S")}.csv'
    )

@app.route('/moderator/logs/export-excel')
@login_required
def export_logs_excel():
    if current_user.role != 'moderator': return redirect(url_for('index'))
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return 'กรุณาติดตั้ง openpyxl: pip install openpyxl', 400
    
    days = request.args.get('days', 7, type=int)
    action = request.args.get('action', '', type=str)
    start_date = datetime.now() - timedelta(days=days)
    query = SystemLog.query.filter(SystemLog.created_at >= start_date)
    if action: query = query.filter_by(action=action)
    logs = query.order_by(SystemLog.created_at.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "System Logs"
    
    headers = ['วันเวลา', 'ผู้ใช้', 'การทำงาน', 'รายละเอียด', 'IP Address']
    ws.append(headers)
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    for log in logs:
        ws.append([
            log.created_at.strftime('%d/%m/%Y %H:%M:%S') if log.created_at else '',
            log.user.username if log.user else 'ลบแล้ว',
            log.action,
            log.detail if log.detail else '',
            log.ip_address if log.ip_address else ''
        ])
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'system_logs_{datetime.now().strftime("%d%m%Y_%H%M%S")}.xlsx'
    )

@app.route('/moderator/logs/export-pdf')
@login_required
def export_logs_pdf():
    if current_user.role != 'moderator': return redirect(url_for('index'))
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        # ⚠️ หมายเหตุ: หากต้องการให้ PDF แสดงภาษาไทยได้ ต้องมีการ Register Font ภาษาไทย (เช่น THSarabunNew) ก่อน
        # from reportlab.pdfbase import pdfmetrics
        # from reportlab.pdfbase.ttfonts import TTFont
        # pdfmetrics.registerFont(TTFont('THSarabunNew', 'path/to/THSarabunNew.ttf'))
    except ImportError:
        return 'กรุณาติดตั้ง reportlab: pip install reportlab', 400
    
    days = request.args.get('days', 7, type=int)
    action = request.args.get('action', '', type=str)
    start_date = datetime.now() - timedelta(days=days)
    query = SystemLog.query.filter(SystemLog.created_at >= start_date)
    if action: query = query.filter_by(action=action)
    logs = query.order_by(SystemLog.created_at.desc()).all()
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'], fontSize=16,
        textColor=colors.HexColor('#1f2937'), spaceAfter=12, alignment=1
        # fontName='THSarabunNew'  # <--- ถ้ามี Font ไทยให้เปิดตรงนี้
    )
    elements.append(Paragraph(f'ประวัติการใช้งานระบบ ({days} วัน)', title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    data = [['วันเวลา', 'ผู้ใช้', 'การทำงาน', 'รายละเอียด', 'IP Address']]
    for log in logs[:100]: 
        data.append([
            log.created_at.strftime('%d/%m/%Y %H:%M') if log.created_at else '',
            log.user.username if log.user else 'ลบแล้ว',
            log.action,
            (log.detail if log.detail else '')[:30],
            log.ip_address if log.ip_address else ''
        ])
    
    table = Table(data, colWidths=[1.3*inch, 1*inch, 1*inch, 2*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), # เปลี่ยนเป็น Font ไทยที่นี่ถ้ามี
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        # ('FONTNAME', (0, 1), (-1, -1), 'THSarabunNew'), # เปลี่ยนเป็น Font ไทยที่นี่ถ้ามี
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'system_logs_{datetime.now().strftime("%d%m%Y_%H%M%S")}.pdf'
    )

# ==================================================
# RUN
# ==================================================
if __name__ == '__main__':
    app.run(debug=True)